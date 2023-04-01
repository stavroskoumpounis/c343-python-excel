import openpyxl
from Course import Course
from Employee import *
from DAO import *
import os

filename = "python_excel.xlsx"

# Create a new Excel workbook with required sheets if it doesn't exist
if not os.path.exists(filename):
    wb = openpyxl.Workbook()
    # Create and set headers for the ListOfTrainees sheet
    trainees_sheet = wb.active
    trainees_sheet.title = "ListOfTrainees"
    headers = ["id", "Name", "Email", "Course", "Background/degree", "WorkExperience"]
    trainees_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_trainee = wb.create_sheet("MappingCourseTrainee")
    headers = ["CourseId", "TraineeId"]
    mapping_course_trainee.append(headers)

    # Create and set headers for the CourseDetails sheet
    course_details_sheet = wb.create_sheet("CourseDetails")
    headers = ["CourseId", "Description"]
    course_details_sheet.append(headers)

    # Create and set headers for the ListOfTrainers sheet
    trainers_sheet = wb.create_sheet("ListOfTrainers")
    headers = ["Email", "Name", "Phone"]
    trainers_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_trainer = wb.create_sheet("MappingCourseTrainer")
    headers = ["CourseId", "TrainerId"]
    mapping_course_trainer.append(headers)

    manager_sheet = wb.create_sheet("ListOfManagers")
    headers = ["Email", "Name", "Phone", "Based"]
    manager_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_manager = wb.create_sheet("MappingCourseManager")
    headers = ["CourseId", "ManagerId"]
    mapping_course_manager.append(headers)

    wb.save(filename)
else:
    wb = openpyxl.load_workbook(filename)
    trainees_sheet = wb['ListOfTrainees']
    mapping_course_trainee = wb['MappingCourseTrainee']
    mapping_course_manager = wb['MappingCourseManager']
    mapping_course_trainer = wb['MappingCourseTrainer']
    course_details_sheet = wb['CourseDetails']
    trainers_sheet = wb['ListOfTrainers']
    manager_sheet = wb['ListOfManagers']

if not os.path.exists("attendance_book.xlsx"):
    attendance_book = openpyxl.Workbook()
    attendance_book.save("attendance_book.xlsx")
else:
    attendance_book = openpyxl.load_workbook("attendance_book.xlsx")


def create_attendance_sheet(date, start, end, course_id):
    attendance_sheet = attendance_book.create_sheet(date)
    headers = ["Date", "StartTime", "EndTime", "CourseId"]
    attendance_sheet.append(headers)
    attendance_sheet.append((date, start, end, course_id))
    attendance_sheet.append(("TraineeId", "Attended"))
    trainee_ids = course_dao.get_all_trainees_for_course(mapping_course_trainee, course_id)
    if trainee_ids:
        for trainee_id in trainee_ids:
            attendance_sheet.append((trainee_id, 1))
    attendance_book.save("attendance_book.xlsx")


trainee1 = Trainee(1, "John Doe", "j.doe@gmail.com", "C343", "CS", 2)
trainee2 = Trainee(2, "Jane Smith", "j.smith@gmail.com", "C345", "Physics", 3)

course1 = Course("C343", "C++ and Python training")
course2 = Course("C345", "Advanced Java Backend")

trainer1 = Trainer("alice.johnson@email.com", "Alice Johnson", "555-123-4567")
trainer2 = Trainer("bob.thompson@email.com", "Bob Thompson", "555-987-6543")

manager = Manager("tomi.ogun@wiley.com", "Tomi O.", "0044-77843123", "EMEA", [course1, course2])

manager_dao.add_managers(manager_sheet, manager)

trainee_dao.add_trainees(trainees_sheet, trainee1, trainee2)
trainer_dao.add_trainers(trainers_sheet, trainer1, trainer2)

wb.save(filename)

trainee3 = trainee_dao.get_trainees_by_id(trainees_sheet, 1)
trainee3[0].name = "John F. Doe"
trainee3[0].email = "Jf.Doe@email.gr"

print(trainee3[0])
trainee_dao.update_trainees(trainees_sheet, trainee3[0])

course_dao.add_course_details(course_details_sheet, course1)
course_dao.add_course_details(course_details_sheet, course2)

course_dao.map_course_trainee(mapping_course_trainee, "C345", trainee1.id)
course_dao.map_course_trainee(mapping_course_trainee, "C343", trainee2.id)

course_dao.map_course_manager(mapping_course_manager, "C343", manager.email)
course_dao.map_course_manager(mapping_course_manager, "C345", manager.email)

course_dao.map_course_trainer(mapping_course_trainer, "C343", trainer1.email)
course_dao.map_course_trainer(mapping_course_trainer, "C345", trainer2.email)

create_attendance_sheet('March30_2023', '9:05', '16:30', course1.course_id)

attendance_dao.mark_absent(attendance_book['March30_2023'], [2])

create_attendance_sheet('March31_2023', '9:05', '16:30', course1.course_id)


wb.save(filename)
