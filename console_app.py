import os
from datetime import datetime
import openpyxl
from DTO import *
from DAO import *

# Your existing code for file creation, reading, and writing here
filename = "python_excel.xlsx"

# Create a new Excel workbook with required sheets if it doesn't exist
if not os.path.exists(filename):
    wb = openpyxl.Workbook()
    # Create and set headers for the ListOfTrainees sheet
    trainees_sheet = wb.active
    trainees_sheet.title = "ListOfTrainees"
    headers = ["id", "Name", "Email", "Course", "Background/degree", "WorkExperience"]
    trainees_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_trainee = wb.create_sheet("MappingCourseTrainee")
    headers = ["CourseId", "TraineeId"]
    mapping_course_trainee.append(headers)

    # Create and set headers for the CourseDetails sheet
    course_details_sheet = wb.create_sheet("CourseDetails")
    headers = ["CourseId", "Description"]
    course_details_sheet.append(headers)

    # Create and set headers for the ListOfTrainers sheet
    trainers_sheet = wb.create_sheet("ListOfTrainers")
    headers = ["Email", "Name", "Phone"]
    trainers_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_trainer = wb.create_sheet("MappingCourseTrainer")
    headers = ["CourseId", "TrainerId"]
    mapping_course_trainer.append(headers)

    manager_sheet = wb.create_sheet("ListOfManagers")
    headers = ["Email", "Name", "Phone", "Based"]
    manager_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_manager = wb.create_sheet("MappingCourseManager")
    headers = ["CourseId", "ManagerId"]
    mapping_course_manager.append(headers)

    wb.save(filename)
else:
    wb = openpyxl.load_workbook(filename)
    trainees_sheet = wb['ListOfTrainees']
    mapping_course_trainee = wb['MappingCourseTrainee']
    mapping_course_manager = wb['MappingCourseManager']
    mapping_course_trainer = wb['MappingCourseTrainer']
    course_details_sheet = wb['CourseDetails']
    trainers_sheet = wb['ListOfTrainers']
    manager_sheet = wb['ListOfManagers']

if not os.path.exists("attendance_book.xlsx"):
    attendance_book = openpyxl.Workbook()
    attendance_book.save("attendance_book.xlsx")
else:
    attendance_book = openpyxl.load_workbook("attendance_book.xlsx")


def create_attendance_sheet(date, start, end, course_id):
    attendance_sheet = attendance_book.create_sheet(date)
    headers = ["Date", "StartTime", "EndTime", "CourseId"]
    attendance_sheet.append(headers)
    attendance_sheet.append((date, start, end, course_id))
    attendance_sheet.append(("TraineeId", "Attended"))
    trainee_ids = course_dao.get_all_trainees_for_course(mapping_course_trainee, course_id)
    if trainee_ids:
        for trainee_id in trainee_ids:
            attendance_sheet.append((trainee_id, 1))
    attendance_book.save("attendance_book.xlsx")


def display_all_trainees(trainees_sheet):
    trainees = trainee_dao.get_all_trainees(trainees_sheet)
    print("\nList of all trainees:")
    print(
        "{:<8}{:<20}{:<25}{:<10}{:<15}{:<15}".format("ID", "Name", "Email", "Course", "Background", "Work Experience"))
    print("-----------------------------------------------------------------------------------------------------")
    for trainee in trainees:
        print("{:<8}{:<20}{:<25}{:<10}{:<15}{:<15}".format(trainee.id, trainee.name, trainee.email, trainee.course,
                                                           trainee.background, trainee.work_exp))
    print("-----------------------------------------------------------------------------------------------------")


def display_trainees_from_attendance(attendance_sheet):
    print("\nList of trainees and attendance status:")
    print("{:<8}{:<20}{:<10}".format("ID", "Name", "Attended"))
    print("--------------------------------------")
    for row in attendance_sheet.iter_rows(min_row=4):
        trainee_id, attended = row[0].value, row[1].value
        trainee = trainee_dao.get_trainees_by_id(trainees_sheet, trainee_id)[0]
        print("{:<8}{:<20}{:<10}".format(trainee.id, trainee.name, attended))
    print("--------------------------------------")


def record_session_attendance():
    date = input("Enter date in format 'May03_2023' (press Enter for current date): ")
    if not date:
        date = datetime.today().strftime('%B%d_%Y')

    start_time = input("Enter start time: ")
    end_time = input("Enter end time: ")

    while True:
        course_id = input("Enter course ID (or type 'exit' to quit): ")
        if course_id.lower() == 'exit':
            return
        trainer_id = course_dao.get_trainer_for_course(mapping_course_trainer, course_id)
        if trainer_id is not None:
            break
        else:
            print("Invalid course ID. Please try again.")

    trainer = trainer_dao.get_trainer_by_email_id(trainers_sheet, trainer_id)
    print(f"Trainer: {trainer.name}")

    create_attendance_sheet(date, start_time, end_time, course_id)
    attendance_sheet = attendance_book[date]

    display_trainees_from_attendance(attendance_sheet)

    absent_ids = input("Enter IDs of absent trainees (comma-separated): ").split(',')
    absent_ids = [int(id.strip()) for id in absent_ids if id.strip()]

    attendance_dao.mark_absent(attendance_sheet, absent_ids)

    # Display the updated list of trainees with their attendance status
    display_trainees_from_attendance(attendance_sheet)

    save = input("Save and submit? (y/n): ")
    if save.lower() == 'y':
        attendance_book.save("attendance_book.xlsx")


def main_menu():
    while True:
        print("\nMain Menu:")
        print("1. Trainees")
        print("2. Trainers[under construction]")
        print("3. Managers[under construction]")
        print("4. Record session attendance")
        print("5. Exit")
        choice = int(input("Enter your choice: "))

        if choice == 1:
            trainee_menu()
        elif choice == 2:
            trainer_menu()
        elif choice == 3:
            manager_menu()
        elif choice == 4:
            record_session_attendance()
        elif choice == 5:
            break
        else:
            print("Invalid choice, try again.")


def add_trainee():
    print("Enter trainee details:")
    id = trainee_dao.get_latest_trainee_id(trainees_sheet) + 1
    name = input("Name: ")
    email = input("Email: ")
    course = input("Course: ")
    background = input("Background/degree: ")
    work_exp = int(input("Work experience: "))
    trainee = Trainee(id, name, email, course, background, work_exp)
    trainee_dao.add_trainees(trainees_sheet, trainee)
    course_dao.map_course_trainee(mapping_course_trainee, course, id)
    wb.save(filename)
    print("Trainee added successfully.")


def update_trainee():
    id = int(input("Enter trainee ID to update: "))
    existing_trainees = trainee_dao.get_trainees_by_id(trainees_sheet, id)
    if not existing_trainees:
        print("Invalid trainee ID.")
    else:
        trainee = existing_trainees[0]
        print("Enter new trainee details (leave blank to keep the old value):")
        name = input(f"Name ({trainee.name}): ")
        email = input(f"Email ({trainee.email}): ")
        course = input(f"Course ({trainee.course}): ")
        background = input(f"Background/degree ({trainee.background}): ")
        work_exp = input(f"Work experience ({trainee.work_exp}): ")

        # Update trainee object with new values if provided
        if name:
            trainee.name = name
        if email:
            trainee.email = email
        if course:
            old_course = trainee.course
            trainee.course = course
            # Update mapping_course_trainee if the course is changed
            course_dao.update_course_trainee_mapping(mapping_course_trainee, old_course, trainee.course, id)
        if background:
            trainee.background = background
        if work_exp:
            trainee.work_exp = int(work_exp)

        trainee_dao.update_trainees(trainees_sheet, trainee)
        wb.save(filename)
        print("Trainee updated successfully.")


def delete_trainee():
    id = int(input("Enter the trainee ID to delete: "))
    trainee_dao.delete_trainees(trainees_sheet, id)
    wb.save(filename)
    print("Trainee deleted successfully.")


def trainee_menu():
    while True:
        print("\nTrainee Menu:")
        print("1. Add trainee")
        print("2. Update trainee")
        print("3. Delete trainee")
        print("4. Display all trainees")
        print("5. Back to main menu")
        choice = int(input("Enter your choice: "))

        if choice == 1:
            add_trainee()
        elif choice == 2:
            update_trainee()
        elif choice == 3:
            delete_trainee()
        elif choice == 4:
            display_all_trainees(trainees_sheet)
        elif choice == 5:
            break
        else:
            print("Invalid choice, try again.")


def trainer_menu():
    pass


def manager_menu():
    pass


if __name__ == "__main__":
    main_menu()
