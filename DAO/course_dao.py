import openpyxl


def map_course_trainee(sheet, course_id, trainee_id):
    sheet.append((course_id, trainee_id))


def map_course_manager(sheet, course_id, email_id):
    sheet.append((course_id, email_id))


def map_course_trainer(sheet, course_id, email_id):
    sheet.append((course_id, email_id))


def add_course_details(sheet, course):
    sheet.append((course.course_id, course.description))


def get_all_trainees_for_course(sheet, course_id):
    trainee_ids = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == course_id:
            trainee_ids.append(row[1].value)
    return trainee_ids


def get_trainer_for_course(sheet, course_id):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == course_id:
            return row[1].value


def update_course_trainee_mapping(sheet, old_course_id, new_course_id, trainee_id):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == old_course_id and row[1].value == trainee_id:
            row[0].value = new_course_id
            break
