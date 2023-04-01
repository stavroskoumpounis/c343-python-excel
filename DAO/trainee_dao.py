import openpyxl

from Employee import Trainee


def add_trainees(sheet, *trainees):
    for trainee in trainees:
        sheet.append((trainee.id, trainee.name, trainee.email, trainee.course, trainee.background, trainee.work_exp))


def delete_trainees(sheet, *trainee_ids):
    for trainee_id in trainee_ids:
        for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
            if row[0] == trainee_id:
                sheet.delete_rows(idx)
                break


def update_trainees(sheet, *trainees):
    for trainee in trainees:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == trainee.id:
                row[1].value = trainee.name
                row[2].value = trainee.email
                row[3].value = trainee.course
                row[4].value = trainee.background
                row[5].value = trainee.work_exp
                break


def get_trainees_by_id(sheet, *trainee_ids):
    trainees = []
    for trainee_id in trainee_ids:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == trainee_id:
                trainee = Trainee(trainee_id, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
                trainees.append(trainee)
                break
    return trainees


def get_all_trainees(sheet):
    trainees = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        trainee = Trainee(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
        trainees.append(trainee)

    return trainees
