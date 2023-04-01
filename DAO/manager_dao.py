import openpyxl

from Employee import Manager


def add_managers(sheet, *managers):
    for manager in managers:
        sheet.append((manager.email, manager.name, manager.phone, manager.based))


def update_managers(sheet, *managers):
    for manager in managers:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == manager.email:
                row[1].value = manager.name
                row[2].value = manager.phone
                row[3].value = manager.based
                break


def delete_managers(sheet, *manager_emails):
    for manager_email in manager_emails:
        for idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
            if row[0] == manager_email:
                sheet.delete_rows(idx)
                break


def get_managers_by_email(sheet, *manager_emails):
    managers = []
    for manager_email in manager_emails:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[1].value == manager_email:
                manager = Manager(row[0].value, manager_email, row[2].value, row[3].value)
                managers.append(manager)
                break
    return managers

