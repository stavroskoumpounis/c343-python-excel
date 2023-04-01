import os
import tkinter as tk
from tkinter import ttk

import openpyxl

from Employee import *
from Course import Course
from DAO import *

# Your existing code for file creation, reading, and writing here
filename = "python_excel.xlsx"

# Create a new Excel workbook with required sheets if it doesn't exist
if not os.path.exists(filename):
    wb = openpyxl.Workbook()
    # Create and set headers for the ListOfTrainees sheet
    trainees_sheet = wb.active
    trainees_sheet.title = "ListOfTrainees"
    headers = ["id", "Name", "Email", "Course", "Background/degree", "WorkExperience"]
    trainees_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_trainee = wb.create_sheet("MappingCourseTrainee")
    headers = ["CourseId", "TraineeId"]
    mapping_course_trainee.append(headers)

    # Create and set headers for the CourseDetails sheet
    course_details_sheet = wb.create_sheet("CourseDetails")
    headers = ["CourseId", "Description"]
    course_details_sheet.append(headers)

    # Create and set headers for the ListOfTrainers sheet
    trainers_sheet = wb.create_sheet("ListOfTrainers")
    headers = ["Email", "Name", "Phone"]
    trainers_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_trainer = wb.create_sheet("MappingCourseTrainer")
    headers = ["CourseId", "TrainerId"]
    mapping_course_trainer.append(headers)

    manager_sheet = wb.create_sheet("ListOfManagers")
    headers = ["Email", "Name", "Phone", "Based"]
    manager_sheet.append(headers)

    # Create and set headers for the MappingCourseTrainee sheet
    mapping_course_manager = wb.create_sheet("MappingCourseManager")
    headers = ["CourseId", "ManagerId"]
    mapping_course_manager.append(headers)

    wb.save(filename)
else:
    wb = openpyxl.load_workbook(filename)
    trainees_sheet = wb['ListOfTrainees']
    mapping_course_trainee = wb['MappingCourseTrainee']
    mapping_course_manager = wb['MappingCourseManager']
    mapping_course_trainer = wb['MappingCourseTrainer']
    course_details_sheet = wb['CourseDetails']
    trainers_sheet = wb['ListOfTrainers']
    manager_sheet = wb['ListOfManagers']

if not os.path.exists("attendance_book.xlsx"):
    attendance_book = openpyxl.Workbook()
    attendance_book.save("attendance_book.xlsx")
else:
    attendance_book = openpyxl.load_workbook("attendance_book.xlsx")


def create_attendance_sheet(date, start, end, course_id):
    attendance_sheet = attendance_book.create_sheet(date)
    headers = ["Date", "StartTime", "EndTime", "CourseId"]
    attendance_sheet.append(headers)
    attendance_sheet.append((date, start, end, course_id))
    attendance_sheet.append(("TraineeId", "Attended"))
    trainee_ids = course_dao.get_all_trainees_for_course(mapping_course_trainee, course_id)
    if trainee_ids:
        for trainee_id in trainee_ids:
            attendance_sheet.append((trainee_id, 1))
    attendance_book.save("attendance_book.xlsx")

# Callback functions for each operation
def add_trainee():
    # Retrieve input data from fields
    trainee_id = int(trainee_id_entry.get())
    name = name_entry.get()
    email = email_entry.get()
    course = course_entry.get()
    background = background_entry.get()
    work_exp = int(work_exp_entry.get())

    # Create a Trainee object and add it to the trainees_sheet
    trainee = Trainee(trainee_id, name, email, course, background, work_exp)
    trainee_dao.add_trainees(trainees_sheet, trainee)
    wb.save(filename)

    # Clear input fields
    trainee_id_entry.delete(0, tk.END)
    name_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    course_entry.delete(0, tk.END)
    background_entry.delete(0, tk.END)
    work_exp_entry.delete(0, tk.END)


# Create the main tkinter window
root = tk.Tk()
root.title("Python Excel")

# Add form fields for trainee information
trainee_id_label = ttk.Label(root, text="Trainee ID:")
trainee_id_entry = ttk.Entry(root)
name_label = ttk.Label(root, text="Name:")
name_entry = ttk.Entry(root)
email_label = ttk.Label(root, text="Email:")
email_entry = ttk.Entry(root)
course_label = ttk.Label(root, text="Course:")
course_entry = ttk.Entry(root)
background_label = ttk.Label(root, text="Background:")
background_entry = ttk.Entry(root)
work_exp_label = ttk.Label(root, text="Work Experience:")
work_exp_entry = ttk.Entry(root)

# Add buttons for performing operations
add_trainee_button = ttk.Button(root, text="Add Trainee", command=add_trainee)

# Grid layout for form fields and buttons
trainee_id_label.grid(column=0, row=0)
trainee_id_entry.grid(column=1, row=0)
name_label.grid(column=0, row=1)
name_entry.grid(column=1, row=1)
email_label.grid(column=0, row=2)
email_entry.grid(column=1, row=2)
course_label.grid(column=0, row=3)
course_entry.grid(column=1, row=3)
background_label.grid(column=0, row=4)
background_entry.grid(column=1, row=4)
work_exp_label.grid(column=0, row=5)
work_exp_entry.grid(column=1, row=5)
add_trainee_button.grid(column=1, row=6)

# Start the main tkinter event loop
root.mainloop()
